import json
import random
# import cehsi
import openpyxl
import requests
import time
from urllib.parse import unquote, quote
import re
from ph import *
from openpyxl import load_workbook, Workbook


# Workbook()方法 不用参数,会新建一个xlsx文件.
wb = openpyxl.Workbook()
# save()方法 一个参数,保存路径,会覆盖.
wb.save('a.xlsx')

def randomwait():
    t = random.randint(1, 3)
    print("等待" + str(t) + "秒")

    time.sleep(t)


def getphone():
    phone = getph()
    return phone


#
# def getid():
#     l = cehsi.outid()
#     name = l[0]
#     id = l[1]
#     return name, id


def gettext(r):
    t = r.text
    p = t[t.index("\\"):t.index("\",")].encode('ascii').decode('unicode_escape')
    print(p)
    return p


def getcode(t):
    code = getcd(t)
    return code


def gettoken(r):
    t = r.text
    p = t[t.index("bPfmSW"):t.index("\"}")].encode('ascii').decode('utf-8')
    print(p)
    return p


def getuserid(r):
    t = r.text
    p = re.findall(r"user_id\":(.+?),", t)
    # a = t.index("user_id")
    # print(a)
    # p = t[a+7:t.index("\"")]
    # p = t
    print(p[0])
    return p[0]


url = 'https://www.chaojijishi.com/h5/#/pages/login/register?invite_code=sbVLkhtlQ'

header1 = {

    'Host': 'www.chaojijishi.com',
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:96.0) Gecko/20100101 Firefox/96.0',
    'Accept': '*/*',
    'Accept-Language': 'zh-CN,zh;q=0.8,zh-TW;q=0.7,zh-HK;q=0.5,en-US;q=0.3,en;q=0.2',
    'Accept-Encoding': 'gzip, deflate, br',
    'content-type': 'application/x-www-form-urlencoded',
    'Content-Length': '233',
    'Origin': 'https://www.chaojijishi.com',
    'Connection': 'keep-alive',
    'Referer': 'https://www.chaojijishi.com/h5/',
    'Sec-Fetch-Dest': 'empty',
    'Sec-Fetch-Mode': 'cors',
    'Sec-Fetch-Site': 'same-origin',
    'TE': 'trailers'

}

session = requests.session()

#

r = session.get(url)
print('开始', r.status_code)
global ck


# print(r.content)

def logn():
    global ck
    url1 = 'https://www.chaojijishi.com/api/mobile/send_verification_codes'
    phone = getphone()
    while phone is None:
        closeph(phone)
        phone = getphone()
    data = {
        'area_code': '86',
        'mobile': phone,
        'login_or_register': 2,
        'XDEBUG_SESSION_START': 'PHPSTORM',
        'timestamp': int(round(time.time() * 1000))
    }
    # 发送验证码
    r = session.post(url1, data=data)
    print('验证码方式', r.status_code)
    p = gettext(r)

    while p != '请求成功':
        phone = getphone()
        while phone is None:
            closeph(phone)
            phone = getphone()
        data['mobile'] = phone
        randomwait()
        r = session.post(url1, data=data)
        print(r.status_code)
        p = gettext(r)

    url2 = 'https://www.chaojijishi.com/api/code/validate_of_mobile_and_code'
    time.sleep(18)
    code = getcode(phone)
    x = 0
    while len(code) < 1:
        time.sleep(5)
        code = getcode(phone)
        x += 1
        print("已尝试", x)
        if x == 6:
            closeph(phone)
            return
    data1 = {
        'mobile': phone,
        'code': code,
        'login_or_register': 2,
        'XDEBUG_SESSION_START': 'PHPSTORM',
        'timestamp': int(round(time.time() * 1000))
    }

    r = session.post(url2, data=data1)
    p = gettext(r)
    while p != '请求成功':
        time.sleep(18)
        code = getcode(phone)
        while len(code) < 1:
            time.sleep(5)
            code = getcode(phone)
            x += 1
            print("尝试", x)
            if x == 6:
                closeph(phone)
                return

        data1['mobile'] = phone
        data1['code'] = code
        r = session.post(url2, data=data1)
        print(r.status_code)
        p = gettext(r)
    closeph(phone)
    return phone, code, r.cookies.get_dict()


def setpassword(phone, code, ck):
    url2 = 'https://www.chaojijishi.com/api/register'
    data2 = {
        'mobile': phone,
        'password': 123456,
        'repeat_pwd': 123456,
        'code': code,
        'login_or_register': 2,
        'invite_code': 'sbVLkhtlQ',
        'XDEBUG_SESSION_START': 'PHPSTORM',
        'timestamp': int(round(time.time() * 1000))
    }
    r = session.post(url2, data=data2, cookies=ck)
    print(r.status_code)
    print('密码设置：')
    gettext(r)


def logn1(phone, ck):
    url3 = 'https://www.chaojijishi.com/api/login'
    data3 = {
        'mobile': phone,
        'password': 123456,
        'XDEBUG_SESSION_START': 'PHPSTORM',
        'timestamp': int(round(time.time() * 1000))
    }
    randomwait()
    r = session.post(url3, data=data3, cookies=ck)
    print(r.status_code)
    print('登录')
    gettext(r)
    return gettoken(r)


# def shiming(ck, tk):
#     url4 = 'https://www.chaojijishi.com/api/user/check_id_card'
#     # url5 = 'https://www.chaojijishi.com/h5/#/pages/subpack1/set/user-id-card-data?type=1'
#     # a = session.get(url5)
#     # tk = quote(tk)
#     # tk = tk.replace('/', '%2F')
#
#     t = getid()
#     print('信息', t)
#     data4 = {
#         'name': t[0],
#         'number': t[1].strip(),
#         'front': 0,
#         'verso': 0,
#         'XDEBUG_SESSION_START': 'PHPSTORM',
#         'timestamp': int(round(time.time() * 1000)),
#         'token': tk
#     }
#     randomwait()
#     r = session.post(url4, data=data4, cookies=ck)
#     res = r.content.decode('utf-8')
#     res = json.loads(res)
#     print(r.status_code)
#     print(res)
#     print('认证')
#     return gettext(r)


def sign():
    yzm = logn()
    if yzm is not None:
        setpassword(yzm[0], yzm[1], yzm[2])
        writeexcle(yzm[0])
        return 1
    else:
        print('失败')
        return 0


i = 0
x = 0
while i < 2:
    try:
        sign()
        print('已完成', i)
        i += 1
    except Exception:
        print('重新来')
        sign()
        print('已完成', i)
        i += 1
发送邮件()
