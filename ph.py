import re
from openpyxl import load_workbook, Workbook
import requests
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.header import Header
import time

token1 = 'dd03c30290d6-239153'
token2 = '425b1a2dc5b6-258930'
apiName = 'MY.431035769'


# 获取token
def gettk():
    url = 'http://api.miyun.pro/api/login?apiName=MY.431035769&password=qq200200'

    r = requests.get(url)

    print(r.content)


# 查询token是否失效
def checktk():
    url1 = 'http://api.miyun.pro/api/get_myinfo?token=' + token2
    r = requests.get(url1)
    print(r.content)


# 获取号码
def getph():
    url2 = 'http://api.miyun.pro/api/get_mobile?token=' + token2 + '&project_id=29756'
    r = requests.get(url2)
    print(r.content)
    t = r.text
    t = re.findall(r"mobile\":\"(.+?)\"", t)
    print('手机号是', t)
    return t


# 获取验证码
def getcd(t):
    url3 = 'http://api.miyun.pro/api/get_message?token=' + token2 + '&project_id=29756&phone_num=' + t[0]
    r = requests.get(url3)
    print(r.content)
    t = r.text
    t = re.findall(r"code\":\"(.+?)\",", t)
    print('验证码是', t)
    return t


# 拉黑号码
def closeph(t):
    url4 = 'http://api.miyun.pro/api/add_blacklist?token=' + token2 + '&project_id=29756&phone_num=' + t[0]
    r = requests.get(url4)
    print(r.content)


# 写入xcel
def writeexcle(phone):
    wb = load_workbook('a.xlsx')
    sheet = wb.active
    print(sheet)
    max_row = sheet.max_row + 1
    print(max_row)
    row_max = 'a' + str(max_row)
    print(row_max)
    sheet[row_max] = str(phone)
    wb.save('a.xlsx')


def 发送邮件():
    # 1. 连接邮箱服务器
    con = smtplib.SMTP_SSL('smtp.163.com', 465)
    youxiangjieshou = 'sch1532694569s@163.com'
    # 2. 登录邮箱
    con.login('sch1532694569s@163.com', 'NVEZDEBNGTNDKCIX')
    # 2. 准备数据
    # 创建邮件对象
    msg = MIMEMultipart()
    # 设置邮件主题
    subject = Header('文件附件发送', 'utf-8').encode()
    msg['Subject'] = subject
    # 设置邮件发送者
    msg['From'] = 'hhh'
    # 设置邮件接受者
    msg['To'] = youxiangjieshou
    # 添加文件附件
    succesful = 'a.xlsx'
    # faile = str((date.today() + timedelta(days=-1))) + '未付款.xlsx'
    file1 = MIMEText(open(succesful, 'rb').read(), 'base64', 'utf-8')
    file1["Content-Disposition"] = 'attachment; filename="' + succesful + '"'
    msg.attach(file1)
    # 中文时使用下面方法
    # if os.path.exists(path + "/" + (str((date.today() + timedelta(days=-1))) + '未付款.xlsx')):
    #     file2 = MIMEText(open(faile, 'rb').read(), 'base64', 'utf-8')
    #     file2.add_header("Content-Disposition", 'attachment', filename=('gbk', '', faile))
    #     msg.attach(file2)
    # 3.发送邮件
    con.sendmail('sch1532694569s@163.com', youxiangjieshou, msg.as_string())
    print("发送成功")
    con.quit()
